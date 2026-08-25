"""从原始资料批量生成开课事实数据。

扫描教学安排表 → 按（学期×课程×专业）建开课记录（含教材映射）→
逐开课导入教学安排表（排课+课程事实）、学期进程表（上课日期）、
学校校历（停课取消+调课+补课占位）、教材资源索引（PPT/大纲/源码）、
人才培养方案（登记+课程类型补齐），并登记源文件。
"""

import re
from pathlib import Path

from openpyxl import load_workbook

import config
import importers
import store
import talent_plan
from resource_indexer import build_resource_index
from template_analyzer import analyze_template

ARRANGEMENT_PATH = config.SOURCE_ROOT / "教学安排表"
PROGRESS_DIR = config.SOURCE_ROOT / "学期进程表"
CALENDAR_DIR = config.SOURCE_ROOT / "学校校历"
TALENT_PLAN_DIR = config.SOURCE_ROOT / "人才培养方案"

# 学期 → 校历文件（按校历内事件日期核实；无对应校历的学期跳过校历导入）
CALENDAR_BY_TERM = {
    "2023-2024-2": "2024年教学历（时间周次）.xls",
    "2024-2025-1": "2024年教学历（时间周次）.xls",
    "2025-2026-2": "2026年教学工作历(1).xls",
    "2026-2027-1": "2026年教学工作历(1).xls",
}


def _normalize_term(raw):
    match = re.fullmatch(r"(\d{4})-(\d{4})学年第([一二])学期", str(raw or "").strip())
    if not match:
        return ""
    semester = "1" if match.group(3) == "一" else "2"
    return f"{match.group(1)}-{match.group(2)}-{semester}"


def _progress_path(term):
    if not term:
        return ""
    for path in sorted(PROGRESS_DIR.glob(f"{term}*.xlsx")):
        return path
    return ""


def _calendar_path(term):
    name = CALENDAR_BY_TERM.get(term)
    return CALENDAR_DIR / name if name else ""


def _template_version(term):
    """按学期匹配模板版本标签（与模板库 version_label 对应）。"""
    if not term:
        return ""
    if term.startswith("2023-2024"):
        return "2023-2024"
    if term in ("2024-2025-1", "2024-2025-2"):
        return term
    if term.startswith("2025-2026") or term.startswith("2026-2027"):
        return "2025-2026"
    return term.rsplit("-", 1)[0]


def scan_arrangements(path=None):
    """扫描教学安排表，返回按（学期×课程×专业）分组的开课组合。"""
    if path is None:
        files = sorted(ARRANGEMENT_PATH.glob("*.xlsx"))
        if not files:
            raise ValueError("教学安排表目录中没有 xlsx 文件。")
        path = files[0]
    workbook = load_workbook(path, data_only=True)
    if "理论教学安排" not in workbook.sheetnames:
        raise ValueError("工作簿中没有“理论教学安排”工作表。")
    groups = {}
    for row in workbook["理论教学安排"].iter_rows(min_row=2, values_only=True):
        term = _normalize_term(row[0])
        course = str(row[4] or "").strip()
        if not term or not course:
            continue
        code_match = re.search(r"\[([^\]]+)\]", course)
        course_code = code_match.group(1).strip() if code_match else ""
        course_name = re.sub(r"^\[[^\]]*\]", "", course).strip()
        course_name = config.COURSE_NAME_ALIASES.get(course_name, course_name)
        class_name = str(row[18] or row[13] or "").strip()
        if not class_name:
            continue
        major = config.resolve_major(class_name)
        key = (term, course_name, major)
        item = groups.setdefault(key, {
            "term": term, "course_name": course_name, "major": major,
            "course_code": course_code, "classes": set(),
        })
        item["classes"].add(class_name)
        if course_code and not item["course_code"]:
            item["course_code"] = course_code
    return path, groups


def _register_source_files(offering_id, term, arrangement_path, textbook_path=""):
    entries = [("教学安排表", str(arrangement_path))]
    progress = _progress_path(term)
    if progress:
        entries.append(("学期进程表", str(progress)))
    calendar = _calendar_path(term)
    if calendar:
        entries.append(("学校校历", str(calendar)))
    if textbook_path:
        entries.append(("教材目录", textbook_path))
    with store.connect() as db:
        for source_type, path in entries:
            existing = db.execute(
                "SELECT id FROM source_files WHERE offering_id=? AND source_type=? ORDER BY id LIMIT 1",
                (offering_id, source_type),
            ).fetchone()
            if existing:
                db.execute(
                    "UPDATE source_files SET source_path=?,notes='批量导入登记（当前路径）' WHERE id=?",
                    (path, existing["id"]),
                )
            else:
                db.execute(
                    "INSERT INTO source_files(offering_id,source_type,source_path,required,notes) "
                    "VALUES (?,?,?,0,'批量导入登记')",
                    (offering_id, source_type, path),
                )
        db.commit()


def _register_templates(offering_id, template_version, offering_kind="普通课程"):
    """按版本从模板库登记模板并完成分析确认（课程标准/授课计划/教学设计+实训资料）。"""
    doc_types = ("授课计划", "实训资料") if offering_kind == "实训课程" else (
        "课程标准", "授课计划", "教学设计", "实训资料")
    registered = []
    with store.connect() as db:
        for doc_type in doc_types:
            if db.execute(
                "SELECT id FROM template_files WHERE offering_id=? AND document_type=?",
                (offering_id, doc_type),
            ).fetchone():
                continue
            lib = store.rows(
                "SELECT * FROM template_library WHERE doc_type=? AND status='已解析' "
                "ORDER BY CASE WHEN version_label=? THEN 0 WHEN version_label='' THEN 2 ELSE 1 END",
                (doc_type, template_version or ""),
            )
            exact = [t for t in lib if t["version_label"] == (template_version or "")]
            year = [t for t in lib if template_version and t["version_label"] and t["version_label"] != template_version
                    and t["version_label"].rsplit("-", 1)[0] == template_version.rsplit("-", 1)[0]]
            generic = [t for t in lib if not t["version_label"]]
            pick = (exact or year or generic or [None])[0]
            if not pick:
                continue
            required = 1 if doc_type in ("课程标准", "授课计划", "教学设计") else 0
            cursor = db.execute(
                "INSERT INTO template_files (offering_id, document_type, template_name, template_path, required, notes) VALUES (?,?,?,?,?,?)",
                (offering_id, doc_type, pick["name"], pick["file_path"], required, "批量登记：模板库自动匹配"),
            )
            db.commit()
            template_file_id = cursor.lastrowid
            try:
                analyze_template(template_file_id)
                db.execute("UPDATE template_analyses SET analysis_status='已确认' WHERE template_file_id=?", (template_file_id,))
                db.commit()
                registered.append(doc_type)
            except Exception:
                db.execute("DELETE FROM template_files WHERE id=?", (template_file_id,))
                db.commit()
    return registered


def _register_talent_plan(offering_id, major, course_name, course_type="", course_code=""):
    """登记人才培养方案到 source_files，并用方案课程表补齐空的课程类型/代码。"""
    plan = talent_plan.get_plan_for_major(major)
    if not plan:
        return False
    with store.connect() as db:
        db.execute(
            "INSERT OR IGNORE INTO source_files(offering_id,source_type,source_path,required,notes) VALUES (?,?,?,0,?)",
            (offering_id, "人才培养方案", plan["source_path"],
             f"{plan.get('cohort') or ''}人才培养方案（按专业匹配）"),
        )
        course_info = talent_plan.find_course(plan, course_name)
        updates, params = [], []
        if not (course_type or "").strip() and course_info and course_info.get("category"):
            updates.append("course_type=?")
            params.append(course_info["category"])
        if not (course_code or "").strip() and course_info and course_info.get("code"):
            updates.append("course_code=?")
            params.append(course_info["code"])
        if updates:
            params.append(offering_id)
            db.execute(f"UPDATE offerings SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
    return True


def bootstrap(force=False):
    """创建缺失的开课记录并导入排课、日期与校历。返回执行摘要。"""
    talent_plan.sync_talent_plans(TALENT_PLAN_DIR)
    path, groups = scan_arrangements()
    created, skipped, imported, failed, no_calendar, no_resources = [], [], [], [], [], []
    for key in sorted(groups):
        term, course_name, major = key
        item = groups[key]
        teaching_class = "；".join(sorted(item["classes"]))
        textbook_path = ""
        textbook_name = ""
        existing = store.rows(
            "SELECT id FROM offerings WHERE term=? AND course_name=? AND major=?",
            (term, course_name, major),
        )
        if existing:
            skipped.append(f"{term}/{course_name}/{major or '未知专业'}")
            offering_id = existing[0]["id"]
            row = store.rows("SELECT textbook_path, textbook_version FROM offerings WHERE id=?", (offering_id,))[0]
            textbook_path = row["textbook_path"]
        else:
            textbook_path, textbook_name = config.resolve_textbook(course_name, term, major)
            textbook_path = str(textbook_path)
            with store.connect() as db:
                cursor = db.execute(
                    """INSERT INTO offerings
                    (term,course_name,course_code,major,teaching_class,textbook_version,textbook_path,
                     schedule_path,notes,offering_kind,total_hours,weekly_hours,credits,template_version)
                    VALUES (?,?,?,?,?,?,?,?,?,?,0,0,0,?)""",
                    (term, course_name, item["course_code"], major, teaching_class,
                     str(textbook_name), str(textbook_path), str(path),
                     "批量导入：教学安排表+教材映射", "普通课程", _template_version(term)),
                )
                db.commit()
                offering_id = cursor.lastrowid
            created.append(f"{term}/{course_name}/{major or '未知专业'}")
        _register_source_files(offering_id, term, path, textbook_path)
        row = store.rows("SELECT template_version, offering_kind FROM offerings WHERE id=?", (offering_id,))[0]
        _register_templates(offering_id, row["template_version"], row["offering_kind"])
        offering_row = store.rows("SELECT course_type, course_code FROM offerings WHERE id=?", (offering_id,))[0]
        _register_talent_plan(offering_id, major, course_name, offering_row["course_type"], offering_row["course_code"])
        if not force and existing:
            continue
        offering = store.rows("SELECT * FROM offerings WHERE id=?", (offering_id,))[0]
        progress = _progress_path(term)
        calendar = _calendar_path(term)
        try:
            calendars = [calendar] if calendar and Path(calendar).exists() else []
            importers.import_schedule_bundle(offering, path, progress, calendars)
            if not calendars:
                no_calendar.append(f"{term}/{course_name}/{major or '未知专业'}")
            try:
                build_resource_index(offering)
            except Exception:
                no_resources.append(f"{term}/{course_name}/{major or '未知专业'}")
            imported.append(f"{term}/{course_name}/{major or '未知专业'}")
        except Exception as exc:
            failed.append(f"{term}/{course_name}/{major or '未知专业'}: {exc}")
    return {
        "arrangement": str(path),
        "created": created,
        "skipped_existing": skipped,
        "imported": imported,
        "failed": failed,
        "no_calendar": no_calendar,
        "no_resources": no_resources,
    }


if __name__ == "__main__":
    import json

    print(json.dumps(bootstrap(), ensure_ascii=False, indent=2))
