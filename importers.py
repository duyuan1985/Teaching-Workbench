import json
import re
import subprocess
import tempfile
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

import store


WEEKDAYS = {"周一", "周二", "周三", "周四", "周五", "周六", "周日"}


def term_label(term):
    match = re.fullmatch(r"(\d{4})-(\d{4})-([12])", term.strip())
    if not match:
        return term
    suffix = "第一学期" if match.group(3) == "1" else "第二学期"
    return f"{match.group(1)}-{match.group(2)}学年{suffix}"


def expand_weeks(expression):
    weeks = []
    for part in str(expression or "").replace("，", ",").split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start, end = (int(value) for value in part.split("-", 1))
            weeks.extend(range(start, end + 1))
        else:
            weeks.append(int(part))
    return sorted(set(weeks))


def parse_slot(value):
    text = str(value or "")
    weekday = next((day for day in WEEKDAYS if day in text), "")
    numbers = [int(number) for number in re.findall(r"\d+", text)]
    return weekday, numbers


def import_teaching_arrangement(offering, path):
    source = Path(path)
    workbook = load_workbook(source, data_only=True)
    if "理论教学安排" not in workbook.sheetnames:
        raise ValueError("工作簿中没有“理论教学安排”工作表。")
    sheet = workbook["理论教学安排"]
    expected_term = term_label(offering["term"])
    matches = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_term = str(row[0] or "").strip()
        course = str(row[4] or "").strip()
        code_match = re.search(r"\[([^\]]+)\]", course)
        row_code = code_match.group(1).strip() if code_match else ""
        code_matches = bool(offering.get("course_code") and row_code == offering["course_code"])
        name_matches = offering["course_name"].lower() in course.lower()
        row_class = str(row[18] or row[13] or "").strip()
        selected_classes = {
            item.strip() for item in re.split(r"[；;、,，\n]+", offering.get("teaching_class", "")) if item.strip()
        }
        class_matches = not selected_classes or row_class in selected_classes
        if row_term == expected_term and (code_matches or name_matches) and class_matches:
            matches.append(row)
    if not matches:
        raise ValueError(f"未找到学期“{expected_term}”、课程“{offering['course_name']}”的排课记录。")

    grouped = {}
    for row in matches:
        class_name = str(row[18] or row[13] or "").strip()
        weeks_text = str(row[19] or "").strip()
        weekday, periods = parse_slot(row[21])
        classroom = str(row[23] or "").strip()
        key = (class_name, weeks_text, weekday, classroom)
        item = grouped.setdefault(key, {"periods": set(), "rows": 0})
        item["periods"].update(periods)
        item["rows"] += 1

    imported = 0
    teacher_name = ""
    first_row = matches[0] if matches else None

    # 从匹配行提取课程级别信息（取第一条匹配行）
    extracted = {}
    if first_row:
        raw_teacher = str(first_row[2] or "").strip()
        if raw_teacher:
            teacher_name = re.sub(r"^\[[^\]]*\]", "", raw_teacher).strip()
        try:
            extracted["credits"] = float(first_row[5] or 0)
        except (TypeError, ValueError):
            pass
        try:
            extracted["total_hours"] = int(float(first_row[6] or 0))
        except (TypeError, ValueError):
            pass
        try:
            extracted["lecture_hours"] = int(float(first_row[7] or 0))
        except (TypeError, ValueError):
            pass
        try:
            extracted["experiment_hours"] = int(float(first_row[8] or 0))
        except (TypeError, ValueError):
            pass
        try:
            extracted["practice_hours"] = int(float(first_row[9] or 0))
        except (TypeError, ValueError):
            pass
        assessment = str(first_row[16] or "").strip()
        if assessment:
            extracted["assessment_type"] = assessment
        category = str(first_row[17] or "").strip()
        if category:
            extracted["course_nature"] = category
        department = str(first_row[1] or "").strip()
        if department:
            extracted["department"] = department
        teaching_mode = str(first_row[24] or "").strip()
        if teaching_mode:
            extracted["teaching_mode"] = teaching_mode

    with store.connect() as db:
        db.execute(
            "DELETE FROM sessions WHERE offering_id=? AND source_note LIKE '教学安排表自动导入:%'",
            (offering["id"],),
        )
        for (class_name, weeks_text, weekday, classroom), item in grouped.items():
            periods = sorted(item["periods"])
            if not periods:
                continue
            period_text = f"{periods[0]}-{periods[-1]}节" if len(periods) > 1 else f"第{periods[0]}节"
            hours = len(periods)
            for week in expand_weeks(weeks_text):
                db.execute(
                    """INSERT INTO sessions
                    (offering_id,class_name,session_type,week_no,lesson_date,weekday,periods,hours,classroom,status,source_note)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        offering["id"], class_name, "正常排课", week, "", weekday, period_text, hours,
                        classroom, "待确认",
                        f"教学安排表自动导入:{source.name}；班级:{class_name}；原周次:{weeks_text}",
                    ),
                )
                imported += 1
        for row in matches:
            class_name = str(row[18] or row[13] or "").strip()
            try:
                enrollment = int(float(row[15] or 0))
            except (TypeError, ValueError):
                enrollment = 0
            if class_name:
                db.execute(
                    """INSERT INTO offering_classes(offering_id,class_name,enrollment_count,source_note,updated_at)
                    VALUES (?,?,?,?,CURRENT_TIMESTAMP)
                    ON CONFLICT(offering_id,class_name) DO UPDATE SET
                    enrollment_count=excluded.enrollment_count,source_note=excluded.source_note,updated_at=CURRENT_TIMESTAMP""",
                    (offering["id"], class_name, enrollment, f"教学安排表选课人数:{source.name}"),
                )
        if teacher_name:
            extracted["teacher_name"] = teacher_name
        if extracted:
            set_clauses = ", ".join(f"{k}=?" for k in extracted)
            db.execute(
                f"UPDATE offerings SET {set_clauses} WHERE id=?",
                list(extracted.values()) + [offering["id"]],
            )
        db.commit()
    return imported, len(matches), len(grouped)


DAY_INDEX = {"周一": 0, "周二": 1, "周三": 2, "周四": 3, "周五": 4, "周六": 5, "周日": 6}
NORMAL_MARKERS = {"∨", "√"}


def _academic_year_for_term(term):
    match = re.fullmatch(r"(\d{4})-(\d{4})-([12])", term.strip())
    if not match:
        raise ValueError("学期格式应为 YYYY-YYYY-1 或 YYYY-YYYY-2。")
    return int(match.group(1)), int(match.group(2)), int(match.group(3))


def _parse_week_start(text, start_year, end_year, semester):
    match = re.search(r"(\d{1,2})\.(\d{1,2})\s*[~～—-]", str(text or ""))
    if not match:
        return None
    month, day = int(match.group(1)), int(match.group(2))
    year = end_year if semester == 2 else start_year
    if semester == 1 and month <= 2:
        year = end_year
    return date(year, month, day)


def import_progress_table(offering, path):
    sessions = store.rows(
        "SELECT * FROM sessions WHERE offering_id=? AND source_note LIKE '教学安排表自动导入:%'",
        (offering["id"],),
    )
    if not sessions:
        raise ValueError("请先导入教学安排表。")
    class_names = set()
    for session in sessions:
        class_name = session.get("class_name", "").strip()
        if class_name:
            class_names.add(class_name)
    start_year, end_year, semester = _academic_year_for_term(offering["term"])
    workbook = load_workbook(path, data_only=True)
    matched_rows = {}
    week_starts = {}
    for sheet in workbook:
        week_row = None
        date_row = None
        for row_no in range(1, min(sheet.max_row, 8) + 1):
            values = [sheet.cell(row_no, col).value for col in range(1, sheet.max_column + 1)]
            if sum(isinstance(value, (int, float)) and 1 <= value <= 30 for value in values) >= 8:
                week_row = row_no
            if sum(bool(re.search(r"\d{1,2}\.\d{1,2}\s*[~～—-]", str(value or ""))) for value in values) >= 8:
                date_row = row_no
        if not week_row or not date_row:
            continue
        for col in range(2, sheet.max_column + 1):
            week_value = sheet.cell(week_row, col).value
            if isinstance(week_value, (int, float)):
                start = _parse_week_start(sheet.cell(date_row, col).value, start_year, end_year, semester)
                if start:
                    week_starts[int(week_value)] = start
        for row_no in range(date_row + 1, sheet.max_row + 1):
            label = str(sheet.cell(row_no, 1).value or "").strip()
            for class_name in class_names:
                if class_name and (class_name in label or label in class_name):
                    matched_rows[class_name] = (sheet, row_no, week_row)

    if not matched_rows:
        raise ValueError("进程表中未找到教学安排表对应的班级。")
    updated = 0
    with store.connect() as db:
        for session in sessions:
            class_name = session.get("class_name", "").strip()
            mapping = matched_rows.get(class_name)
            # 重新导入进程表只更新日期和状态，保留教学安排表生成的原始周次。
            week = int(session["week_no"] or 0)
            start = week_starts.get(week)
            if not mapping or not start or session["weekday"] not in DAY_INDEX:
                continue
            sheet, row_no, week_row = mapping
            week_col = next(
                (col for col in range(2, sheet.max_column + 1) if sheet.cell(week_row, col).value == week),
                None,
            )
            marker = str(sheet.cell(row_no, week_col).value or "").strip() if week_col else ""
            lesson_date = start + timedelta(days=DAY_INDEX[session["weekday"]])
            status = "已确认" if not marker or marker in NORMAL_MARKERS else "待确认"
            note = session["source_note"] + f"；进程表:{Path(path).name}；周状态:{marker or '空白'}"
            db.execute(
                "UPDATE sessions SET lesson_date=?,status=?,source_note=? WHERE id=?",
                (lesson_date.isoformat(), status, note, session["id"]),
            )
            updated += 1
        db.commit()
    return updated, len(matched_rows), len(week_starts)


EVENT_NAMES = ("元旦", "清明节", "劳动节", "端午节", "中秋节", "国庆节", "运动会")


def _calendar_rows(path):
    source_path = Path(path)
    if source_path.suffix.lower() == ".xls":
        helper = Path(__file__).with_name("read_legacy_xls.ps1")
        with tempfile.TemporaryDirectory(prefix="teaching-calendar-") as temp_dir:
            output_path = Path(temp_dir) / "calendar.json"
            result = subprocess.run(
                [
                    "powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass",
                    "-File", str(helper), "-InputPath", str(source_path),
                    "-OutputPath", str(output_path),
                ],
                capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=90,
            )
            if result.returncode != 0 or not output_path.exists():
                detail = (result.stderr or result.stdout or "Excel读取失败").strip()
                raise ValueError(f"旧版校历读取失败：{detail}")
            sheets = json.loads(output_path.read_text(encoding="utf-8"))
            return [row for sheet in sheets for row in sheet.get("rows", [])]
    workbook = load_workbook(path, data_only=True)
    return [list(row) for sheet in workbook for row in sheet.iter_rows(values_only=True)]


def import_calendar_week_dates(offering, path):
    rows = _calendar_rows(path)
    expected_title = term_label(offering["term"])
    title_col = None
    title_row = None
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            if expected_title in str(value or ""):
                title_row, title_col = row_index, col_index
                break
        if title_col is not None:
            break
    if title_col is None:
        raise ValueError(f"校历中未找到“{expected_title}”。")

    week_col = None
    for row in rows[title_row:title_row + 8]:
        for col_index in range(title_col, min(len(row), title_col + 5)):
            if str(row[col_index] or "").strip() == "周次":
                week_col = col_index
                break
        if week_col is not None:
            break
    if week_col is None:
        raise ValueError("校历中未找到周次列。")

    month_col = title_col
    monday_col = week_col + 1
    current_year = None
    current_month = None
    week_starts = {}
    for row in rows[title_row:]:
        month_text = str(row[month_col] or "").strip() if month_col < len(row) else ""
        month_match = re.fullmatch(r"(\d{4})\.(\d{1,2})", month_text)
        if month_match:
            current_year, current_month = (int(value) for value in month_match.groups())
        week_text = str(row[week_col] or "").strip() if week_col < len(row) else ""
        monday_text = str(row[monday_col] or "").strip() if monday_col < len(row) else ""
        if not (week_text.isdigit() and monday_text.isdigit() and current_year and current_month):
            continue
        week = int(week_text)
        if week not in week_starts:
            try:
                week_starts[week] = date(current_year, current_month, int(monday_text))
            except ValueError:
                continue

    updated = 0
    with store.connect() as db:
        sessions = [dict(row) for row in db.execute(
            "SELECT * FROM sessions WHERE offering_id=? AND source_note LIKE '教学安排表自动导入:%'",
            (offering["id"],),
        )]
        for session in sessions:
            start = week_starts.get(session["week_no"])
            if not start or session["weekday"] not in DAY_INDEX:
                continue
            lesson_date = start + timedelta(days=DAY_INDEX[session["weekday"]])
            note = session["source_note"] + f"；校历周次:{Path(path).name}"
            db.execute(
                "UPDATE sessions SET lesson_date=?,status='已确认',source_note=? WHERE id=?",
                (lesson_date.isoformat(), note, session["id"]),
            )
            updated += 1
        db.commit()
    return updated, len(week_starts)


def _event_ranges(text, year):
    ranges = []
    for name in EVENT_NAMES:
        if name not in text:
            continue
        found = False
        patterns = [
            rf"{re.escape(name)}[^。；]*?(\d{{1,2}})月(\d{{1,2}})日至(\d{{1,2}})月(\d{{1,2}})日",
            rf"{re.escape(name)}[^。；]*?(\d{{1,2}})月(\d{{1,2}})日至(\d{{1,2}})日",
            rf"{re.escape(name)}[^。；]*?(\d{{1,2}})月(\d{{1,2}})[-—](\d{{1,2}})日",
            rf"{re.escape(name)}[^。；]*?(\d{{1,2}})月(\d{{1,2}})日[、,，](\d{{1,2}})日",
        ]
        for index, pattern in enumerate(patterns):
            match = re.search(pattern, text)
            if not match:
                continue
            values = [int(value) for value in match.groups()]
            if index == 0:
                start_month, start_day, end_month, end_day = values
            else:
                start_month, start_day, end_day = values
                end_month = start_month
            ranges.append((name, date(year, start_month, start_day), date(year, end_month, end_day)))
            found = True
            break
        if not found and name == "运动会":
            match = re.search(r"(\d{1,2})月(\d{1,2})日[、,，](\d{1,2})日[^。；]*运动会", text)
            if match:
                month, start_day, end_day = (int(value) for value in match.groups())
                ranges.append((name, date(year, month, start_day), date(year, month, end_day)))
    return ranges


def import_school_calendar(offering, path):
    start_year, end_year, semester = _academic_year_for_term(offering["term"])
    calendar_year = end_year if semester == 2 else start_year
    texts = []
    for row in _calendar_rows(path):
        for value in row:
            text = str(value or "").strip()
            if text and any(name in text for name in EVENT_NAMES):
                texts.append(text)
    events = []
    replacements = []
    for text in texts:
        for name, start, end in _event_ranges(text, calendar_year):
            if semester == 2 and not 2 <= start.month <= 8:
                continue
            if semester == 1 and 2 < start.month < 8:
                continue
            event_pos = text.find(name)
            next_positions = [text.find(other, event_pos + len(name)) for other in EVENT_NAMES]
            next_positions = [position for position in next_positions if position > event_pos]
            segment = text[event_pos:min(next_positions)] if next_positions else text[event_pos:]
            replacement_dates = re.findall(
                r"(\d{1,2})月(\d{1,2})日\s*[（(](?:星期|周)[一二三四五六日][）)]",
                segment,
            )
            replacement_targets = re.findall(
                r"第(\d+)周(周[一二三四五六日])(?:的课)?",
                segment,
            )
            for week, first_day, second_day in re.findall(
                r"第(\d+)周(周[一二三四五六日])\s*[、,，和及]\s*(周[一二三四五六日])",
                segment,
            ):
                continuation = (week, second_day)
                if continuation not in replacement_targets:
                    replacement_targets.append(continuation)
            replacements = list(zip(replacement_dates, replacement_targets))
            if not replacements:
                replacements = [(None, None)]
            for replacement_date_parts, replacement_target in replacements:
                replacement_date = ""
                target_week = None
                target_weekday = ""
                if replacement_date_parts and replacement_target:
                    month, day = replacement_date_parts
                    target_week, target_weekday = replacement_target
                    replacement_date = date(calendar_year, int(month), int(day)).isoformat()
                    target_week = int(target_week)
                events.append({
                    "name": name,
                    "start": start,
                    "end": end,
                    "replacement_date": replacement_date,
                    "target_week": target_week,
                    "target_weekday": target_weekday,
                    "text": text,
                })

    unique = {}
    for event in events:
        key = (
            event["name"], event["start"], event["end"],
            event["replacement_date"], event["target_week"], event["target_weekday"],
        )
        current = unique.get(key)
        if current is None or event["replacement_date"]:
            unique[key] = event
    events = list(unique.values())
    source = Path(path)
    with store.connect() as db:
        db.execute(
            "DELETE FROM calendar_events WHERE offering_id=? AND source_note LIKE '校历自动导入:%'",
            (offering["id"],),
        )
        db.execute(
            "DELETE FROM sessions WHERE offering_id=? AND source_note LIKE '校历自动生成:%'",
            (offering["id"],),
        )
        sessions = [dict(row) for row in db.execute(
            "SELECT * FROM sessions WHERE offering_id=? AND lesson_date<>''",
            (offering["id"],),
        )]
        cancelled = 0
        generated = 0
        for event in events:
            db.execute(
                """INSERT INTO calendar_events
                (offering_id,event_name,start_date,end_date,suspends_classes,replacement_date,target_week_no,target_weekday,source_note)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (
                    offering["id"], event["name"], event["start"].isoformat(), event["end"].isoformat(),
                    1, event["replacement_date"], event["target_week"], event["target_weekday"],
                    f"校历自动导入:{source.name}；原文:{event['text']}",
                ),
            )
            for session in sessions:
                session_date = date.fromisoformat(session["lesson_date"])
                if event["start"] <= session_date <= event["end"]:
                    db.execute(
                        "UPDATE sessions SET status='已取消',source_note=source_note||? "
                        "WHERE id=? AND source_note NOT LIKE ?",
                        (
                            f"；校历停课:{event['name']}", session["id"],
                            f"%校历停课:{event['name']}%",
                        ),
                    )
                    cancelled += 1
            if event["replacement_date"] and event["target_week"] and event["target_weekday"]:
                targets = [
                    session for session in sessions
                    if session["week_no"] == event["target_week"] and session["weekday"] == event["target_weekday"]
                ]
                for target in targets:
                    db.execute(
                        """INSERT INTO sessions
                        (offering_id,class_name,session_type,week_no,lesson_date,weekday,periods,hours,classroom,status,source_note)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            offering["id"], target.get("class_name", ""), "调课", event["target_week"],
                            event["replacement_date"],
                            tuple(DAY_INDEX)[date.fromisoformat(event["replacement_date"]).weekday()],
                            target["periods"], target["hours"], target["classroom"], "已确认",
                            f"校历自动生成:{source.name}；补第{event['target_week']}周{event['target_weekday']}课程",
                        ),
                    )
                    generated += 1
        db.commit()
    return len(events), cancelled, generated


def create_makeup_placeholders(offering):
    """Create editable makeup requirements only after normal sessions are resolved."""
    expected = int(offering["total_hours"])
    block = max(1, int(offering.get("weekly_hours") or 4))
    with store.connect() as db:
        db.execute(
            "DELETE FROM sessions WHERE offering_id=? AND (source_note LIKE '系统待补课:%' OR source_note LIKE '系统第18周补课:%')",
            (offering["id"],),
        )
        rows = [dict(row) for row in db.execute(
            "SELECT * FROM sessions WHERE offering_id=?",
            (offering["id"],),
        )]
        class_names = sorted({row.get("class_name", "").strip() for row in rows if row.get("class_name", "").strip()})
        created = 0
        for class_name in class_names:
            class_rows = [row for row in rows if row.get("class_name", "").strip() == class_name]
            if any(row["session_type"] == "正常排课" and row["status"] == "待确认" for row in class_rows):
                continue
            confirmed = sum(
                int(row["hours"]) for row in class_rows
                if row["status"] == "已确认" and row["session_type"] != "停课"
            )
            remaining = expected - confirmed
            while remaining > 0:
                hours = min(block, remaining)
                db.execute(
                    """INSERT INTO sessions
                    (offering_id,class_name,session_type,week_no,lesson_date,weekday,periods,hours,classroom,status,source_note)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        offering["id"], class_name, "补课", 18, "", "", "", hours, "", "已确认",
                        f"系统第18周补课:依据课程总学时{expected}与已确认排课{confirmed}的差额自动生成",
                    ),
                )
                remaining -= hours
                created += 1
        db.commit()
    return created


def import_schedule_bundle(offering, arrangement_path, progress_path="", calendar_paths=()):
    """Import schedule facts in the same order for CLI and UI workflows."""
    calendar_paths = list(calendar_paths or ())
    has_progress = bool(progress_path)
    with store.connect() as db:
        db.execute(
            "DELETE FROM sessions WHERE offering_id=? AND (source_note LIKE '教学安排表自动导入:%' OR source_note LIKE '校历自动生成:%')",
            (offering["id"],),
        )
        db.execute(
            "DELETE FROM calendar_events WHERE offering_id=? AND source_note LIKE '校历自动导入:%'",
            (offering["id"],),
        )
        db.commit()
    arrangement = import_teaching_arrangement(offering, arrangement_path)
    progress = import_progress_table(offering, progress_path) if has_progress else None
    calendars = []
    calendar_errors = []
    calendar_week_dates = []
    for path in calendar_paths:
        try:
            if not has_progress:
                calendar_week_dates.append(import_calendar_week_dates(offering, path))
            calendars.append(import_school_calendar(offering, path))
        except Exception as error:
            calendar_errors.append(f"{Path(path).name}: {error}")
    placeholders = create_makeup_placeholders(offering) if calendars and not calendar_errors else 0
    return {
        "arrangement": arrangement,
        "progress": progress,
        "calendars": calendars,
        "calendar_errors": calendar_errors,
        "calendar_week_dates": calendar_week_dates,
        "progress_missing": not has_progress,
        "makeup_placeholders": placeholders,
    }


def rebuild_schedule(offering):
    sources = store.rows(
        "SELECT * FROM source_files WHERE offering_id=? ORDER BY id",
        (offering["id"],),
    )
    by_type = {}
    for source in sources:
        path = source["source_path"]
        if Path(path).exists():
            by_type.setdefault(source["source_type"], []).append(path)
    if not by_type.get("教学安排表"):
        raise ValueError("源文件中缺少可用的教学安排表。")
    return import_schedule_bundle(
        offering,
        by_type["教学安排表"][0],
        (by_type.get("学期进程表") or [""])[0],
        by_type.get("学校校历", []),
    )
