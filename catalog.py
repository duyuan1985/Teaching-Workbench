import re
from pathlib import Path

from openpyxl import load_workbook


def term_label(term):
    match = re.fullmatch(r"(\d{4})-(\d{4})-([12])", term)
    if not match:
        return term
    return f"{match.group(1)}-{match.group(2)}学年{'第一学期' if match.group(3)=='1' else '第二学期'}"


def major_from_class(class_name):
    text = str(class_name or "")
    if "电商" in text or "农商" in text:
        return "农村电子商务"
    if "全媒体" in text:
        return "全媒体广告策划与营销"
    if "营销" in text:
        return "市场营销"
    return "其他专业"


def catalog_from_arrangement(path, term):
    workbook = load_workbook(Path(path), data_only=True)
    sheet = workbook["理论教学安排"] if "理论教学安排" in workbook.sheetnames else workbook.active
    expected = term_label(term)
    grouped = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() != expected:
            continue
        raw_course = str(row[4] or "").strip()
        code_match = re.search(r"\[([^\]]+)\]", raw_course)
        code = code_match.group(1) if code_match else ""
        course_name = re.sub(r"^\[[^\]]+\]", "", raw_course).strip()
        class_name = str(row[18] or row[13] or "").strip()
        slot = str(row[21] or "")
        periods = re.findall(r"\d+", slot)
        major = major_from_class(class_name)
        key = (course_name, code, major)
        item = grouped.setdefault(key, {
            "course_name": course_name, "course_code": code, "major": major,
            "class_names": set(), "class_counts": {}, "credits": row[5] or 0, "total_hours": row[6] or 0,
            "weekly_hours": 0, "slots": set(),
            "course_nature": "选修课" if str(row[17] or "").strip() == "限选课" else "必修课" if str(row[17] or "").strip() == "必修课" else str(row[17] or "").strip(),
            "assessment_type": "期末考核",
            "assessment_method": "实操",
        })
        item["class_names"].add(class_name)
        try:
            item["class_counts"][class_name] = max(item["class_counts"].get(class_name, 0), int(float(row[15] or 0)))
        except (TypeError, ValueError):
            pass
        item["slots"].add(slot)
        item["weekly_hours"] = max(item["weekly_hours"], len(periods))
    result = []
    for item in grouped.values():
        item["weekly_hours"] = max(1, item["weekly_hours"])
        item["class_name"] = "；".join(sorted(item.pop("class_names")))
        item.pop("slots", None)
        result.append(item)
    return sorted(result, key=lambda x: (x["course_name"], x["class_name"]))
