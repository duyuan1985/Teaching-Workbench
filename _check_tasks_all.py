import store

# 查看tasks表完整内容
all_tasks = store.rows("SELECT offering_id, COUNT(*) as cnt FROM tasks GROUP BY offering_id ORDER BY offering_id")
print("=== tasks表按课程统计 ===")
for t in all_tasks:
    offering = store.rows("SELECT course_name, term FROM offerings WHERE id=?", [t['offering_id']])
    name = offering[0]['course_name'] if offering else '?'
    term = offering[0]['term'] if offering else '?'
    print(f"  offering_id={t['offering_id']}: {name} ({term}) → {t['cnt']}个任务")

# 检查是否有其他表存储任务
tables = store.rows("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE '%task%' OR name LIKE '%assignment%'")
print(f"\n任务相关表: {[t['name'] for t in tables]}")

# 检查assignments表
try:
    assignments = store.rows("SELECT offering_id, COUNT(*) as cnt FROM assignments GROUP BY offering_id ORDER BY offering_id")
    print(f"\n=== assignments表 ===")
    for a in assignments:
        print(f"  offering_id={a['offering_id']}: {a['cnt']}条")
except Exception as e:
    print(f"assignments表查询失败: {e}")

# 检查教学安排表原始数据
import os
schedule_path = r'原始资料\教学安排表\教学安排表20260816115742.xlsx'
if os.path.exists(schedule_path):
    from openpyxl import load_workbook
    wb = load_workbook(schedule_path, read_only=True)
    print(f"\n=== 教学安排表 ===")
    print(f"工作表: {wb.sheetnames}")
    for sn in wb.sheetnames[:3]:
        ws = wb[sn]
        print(f"\n  [{sn}] {ws.max_row}行 × {ws.max_column}列")
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
            print(f"    R{ri+1}: {[str(c)[:25] if c else '' for c in row[:8]]}")
    wb.close()
