from pathlib import Path
from openpyxl import load_workbook

print("=== 教材目录 ===")
for d in sorted(Path("原始资料/教材").iterdir()):
    if d.is_dir():
        packs = [p.name for p in d.iterdir() if p.is_dir()]
        print(f"{d.name}: {packs}")

wb = load_workbook("原始资料/教学安排表/教学安排表20260816115742.xlsx", data_only=True)
print("\n=== 工作表 ===", wb.sheetnames)
sheet = wb["理论教学安排"]
print("\n=== 表头与前2行 ===")
for row in sheet.iter_rows(min_row=1, max_row=3, values_only=True):
    print([str(v)[:14] if v is not None else "" for v in row[:26]])

print("\n=== 全部学期×课程×班级组合（去重）===")
combos = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
    term = str(row[0] or "").strip()
    course = str(row[4] or "").strip()
    cls = str(row[18] or row[13] or "").strip()
    if term and course:
        combos.setdefault((term, course), set()).add(cls)
for (term, course), classes in sorted(combos.items()):
    print(f"{term} | {course[:34]} | {sorted(classes)}")
