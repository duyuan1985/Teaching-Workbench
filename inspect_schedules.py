from openpyxl import load_workbook
from pathlib import Path
import sys

sys.stdout.reconfigure(encoding='utf-8')

FILES = [
    Path(r'E:\开发\AIGC\教学安排表20260816115742.xlsx'),
    Path(r'E:\工作\90-历史学期归档\2023-2024-2\课表与进程表\2023-2024-2进程表（经贸系）.xlsx'),
    Path(r'E:\开发\AIGC\2026年教学工作历(1).xlsx'),
]

for path in FILES:
    print(f'FILE {path}')
    workbook = load_workbook(path, data_only=True)
    for sheet in workbook:
        print(f'SHEET {sheet.title} {sheet.max_row} {sheet.max_column}')
        for index, row in enumerate(sheet.iter_rows(values_only=True), 1):
            values = [str(value) for value in row if value not in (None, '')]
            if values:
                print(index, ' | '.join(values[:40]))
