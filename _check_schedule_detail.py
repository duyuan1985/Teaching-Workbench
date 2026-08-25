from openpyxl import load_workbook

schedule_path = r'原始资料\教学安排表\教学安排表20260816115742.xlsx'
wb = load_workbook(schedule_path, read_only=True)
ws = wb['理论教学安排']

print(f'理论教学安排: {ws.max_row}行 × {ws.max_column}列\n')

# 打印前20行的全部列
for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
    vals = [str(c)[:20] if c else '' for c in row[:15]]
    print(f'R{ri+1}: {vals}')

print('\n...\n')

# 找出所有不同的课程
courses = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[4]:
        courses.add(str(row[4]))
print(f'共{len(courses)}个不同课程:')
for c in sorted(courses):
    print(f'  {c}')

wb.close()
